import cv2 as cv
import cvzone
import os
import pickle
import face_recognition as fr
import numpy as np
import datetime
from openpyxl import load_workbook


cap = cv.VideoCapture(0)
cap.set(3, 640)
cap.set(4, 480)

imgBackground = cv.imread("Resources/background.png")

folderModePath = "Resources/Modes"
modePathList = os.listdir(folderModePath)
imgModeList = []

for path in modePathList:
    imgModeList.append(cv.imread(os.path.join(folderModePath, path)))


print("\n\nLOADING ENCODE FILE START ...")
file = open("EncodeFile.p", "rb")
encodeListKnownWithIds = pickle.load(file)
file.close()
encodeListKnown, personIds = encodeListKnownWithIds
# print(personIds)
print("ENCODE FILE LOADED\n\n")

today = datetime.date.today()
filename = "attendance.xlsx"
workbook = load_workbook(filename)
sheet = workbook.active

# پیدا کردن ستون مربوط به تاریخ امروز (اگر وجود داشته باشد)
column_index = None
for col in range(1, sheet.max_column + 1):
    column_date = sheet.cell(row=1, column=col).value
    if isinstance(column_date, datetime.datetime) and column_date.date() == today:
        column_index = col
        break

# اگر ستون برای تاریخ امروز وجود نداشت، ستون جدیدی بسازید و با today نام گذاری کنید
if column_index is None:
    column_index = sheet.max_column + 1
    sheet.cell(row=1, column=column_index).value = today

isNameWrittenDic = {}
# color = (17, 0, 153)

while True:
    success, img = cap.read()

    imgS = cv.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv.cvtColor(imgS, cv.COLOR_BGR2RGB)

    faceCurFrame = fr.face_locations(imgS)
    encodeCurFrame = fr.face_encodings(imgS, faceCurFrame)

    imgBackground[162:162+480, 55:55+640] = img
    imgBackground[44:44+633, 808:808+414] = imgModeList[0]

    for encoFace, faceLoc in zip(encodeCurFrame, faceCurFrame):
        matches = fr.compare_faces(encodeListKnown, encoFace)
        faceDis = fr.face_distance(encodeListKnown, encoFace)

        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            personID = personIds[matchIndex]
            top, right, bottom, left = faceLoc
            top, right, bottom, left = top*4, right*4, bottom*4, left * 4
            bbox = 55+left, 162+top, right-left, bottom-top
            imgBackground = cvzone.cornerRect(imgBackground, bbox, rt=0)
            # cv.rectangle(
            #     imgBackground, (bbox[0], bbox[1]), (bbox[0] + bbox[2], bbox[1] + bbox[3]), color, 2)
            imgBackground[44:44+633, 808:808+414] = imgModeList[1]
            cv.putText(imgBackground, personID,
                       (bbox[0]+10, bbox[1]-10), cv.FONT_HERSHEY_DUPLEX, 0.9, (17, 0, 153), 2)
            if personID not in isNameWrittenDic or not isNameWrittenDic[personID]:
                name = personID
                print(personIds[matchIndex])
                current_time = datetime.datetime.now().strftime("%H:%M:%S")
                # پیدا کردن سطر مربوط به نام شخص
                row_index = None
                for row in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=1).value
                    if cell_value == name:
                        row_index = row
                        break

                # اگر نام شخص پیدا شد، ساعت را در ستون مربوطه ثبت کنید
                if row_index is not None:
                    column_index = None
                    for col in range(1, sheet.max_column + 1):
                        column_date = sheet.cell(row=1, column=col).value
                        if isinstance(column_date, datetime.datetime) and column_date.date() == today:
                            column_index = col
                            break
                    # اگر ستون برای تاریخ امروز وجود نداشت، ستون جدیدی بسازید و با today نام گذاری کنید
                    if column_index is None:
                        column_index = sheet.max_column + 1
                        sheet.cell(row=1, column=column_index).value = today

                    # ثبت ساعت در ستون مربوطه
                    sheet.cell(row=row_index,
                               column=column_index).value = current_time
                isNameWrittenDic[personID] = True

        else:
            imgBackground[44:44+633, 808:808+414] = imgModeList[2]

    if success:
        height, width, _ = img.shape
        if height > 0 and width > 0:
            cv.imshow("Face Attendance", imgBackground)
            if cv.waitKey(1) & 0xFF == ord('q'):
                break
        else:
            print("Invalid image size.")
            break
    else:
        print("Failed to capture frame.")
        break
# ذخیره فایل اکسل
workbook.save(filename)

# بستن فایل اکسل
workbook.close()

cv.destroyAllWindows()

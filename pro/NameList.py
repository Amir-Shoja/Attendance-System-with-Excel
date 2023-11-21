from openpyxl import Workbook
import pickle


file = open("EncodeFile.p", "rb")
encodeListKnownWithIds = pickle.load(file)
file.close()
encodeListKnown, personIds = encodeListKnownWithIds

# ایجاد یک کتاب کار
workbook = Workbook()

# ایجاد یک ورک‌شیت
worksheet = workbook.active

# تنظیم نام و ستون‌ها
worksheet.title = "attendance"
worksheet.cell(row=1, column=1, value="Name")
# worksheet.cell(row=1, column=2, value="Date & Time")

# وارد کردن اسامی در ستون اول
for i, name in enumerate(personIds, start=2):
    worksheet.cell(row=i, column=1, value=name)

# ذخیره کردن فایل اکسل
workbook.save('attendance.xlsx')
